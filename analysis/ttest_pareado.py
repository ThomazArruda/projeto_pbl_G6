"""
Análise de Teste T Pareado - Comparação de Pernas Esquerda vs Direita
Script para realizar testes t pareados comparando perna esquerda (parética) 
com perna direita (controle) para ângulo, EMG e ECG - Sessões 19-23
"""

import sys
import io

# Configurar encoding para UTF-8 no Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

import sqlite3
import json
import pandas as pd
import numpy as np
from scipy import stats
from pathlib import Path
from datetime import datetime


class PairedTTestAnalyzer:
    """Classe para análise de testes t pareados entre pernas"""
    
    def __init__(self, db_path="../backend/clinic.db"):
        """
        Inicializa o analisador com caminho para o banco de dados
        
        Args:
            db_path: Caminho para o arquivo clinic.db
        """
        self.db_path = Path(db_path)
        self.conn = None
        self.sessions_data = {}
        
        # Armazenar deltas por variável
        self.deltas = {
            'angulo': {'esq': [], 'dir': []},
            'emg': {'esq': [], 'dir': []},
            'ecg': {'esq': [], 'dir': []}
        }
        
        self.results = {}
        self.session_ids = [19, 20, 21, 22, 23]
        
    def connect_db(self):
        """Conecta ao banco de dados SQLite"""
        try:
            self.conn = sqlite3.connect(str(self.db_path))
            print(f"✓ Conectado ao banco de dados: {self.db_path}")
        except Exception as e:
            print(f"✗ Erro ao conectar ao banco de dados: {e}")
            raise
    
    def close_db(self):
        """Fecha a conexão com o banco de dados"""
        if self.conn:
            self.conn.close()
            print("✓ Conexão com banco de dados fechada")
    
    def extract_session_data(self):
        """
        Extrai dados das sessões especificadas
        """
        print(f"\n{'='*80}")
        print(f"EXTRAÇÃO DE DADOS - Sessões {self.session_ids[0]} a {self.session_ids[-1]}")
        print(f"{'='*80}")
        
        query = f"""
            SELECT id, timestamp, duration_seconds, raw_data_blob 
            FROM sessions 
            WHERE id IN ({','.join(map(str, self.session_ids))})
            ORDER BY id
        """
        
        try:
            df = pd.read_sql_query(query, self.conn)
            
            if df.empty:
                print(f"✗ Nenhuma sessão encontrada para IDs: {self.session_ids}")
                return False
            
            print(f"✓ {len(df)} sessão(ões) encontrada(s)\n")
            
            for idx, row in df.iterrows():
                session_id = row['id']
                try:
                    raw_data = json.loads(row['raw_data_blob'])
                    self.sessions_data[session_id] = {
                        'timestamp': row['timestamp'],
                        'duration': row['duration_seconds'],
                        'raw_data': raw_data
                    }
                    print(f"  Sessão {session_id}: ✓ Dados extraídos")
                except json.JSONDecodeError as e:
                    print(f"  Sessão {session_id}: ✗ Erro ao decodificar JSON: {e}")
            
            return True
            
        except Exception as e:
            print(f"✗ Erro ao extrair dados das sessões: {e}")
            raise
    
    def calculate_deltas(self):
        """
        Calcula deltas (max - min) para ângulo, EMG e ECG
        """
        print(f"\n{'='*80}")
        print(f"CÁLCULO DE DELTAS (MÁXIMO - MÍNIMO)")
        print(f"{'='*80}\n")
        
        for session_id, data in sorted(self.sessions_data.items()):
            try:
                raw_data = data['raw_data']
                
                # Extrair dados de ambas as pernas
                angles_esq = []
                angles_dir = []
                emg_esq = []
                emg_dir = []
                ecg_esq = []
                ecg_dir = []
                
                if isinstance(raw_data, list):
                    for point in raw_data:
                        if isinstance(point, dict):
                            # Ângulos
                            if 'ESQ_angle' in point:
                                angles_esq.append(point['ESQ_angle'])
                            if 'DIR_angle' in point:
                                angles_dir.append(point['DIR_angle'])
                            
                            # EMG
                            if 'ESQ_emg' in point:
                                emg_esq.append(point['ESQ_emg'])
                            if 'DIR_emg' in point:
                                emg_dir.append(point['DIR_emg'])
                            
                            # ECG
                            if 'ESQ_ecg' in point:
                                ecg_esq.append(point['ESQ_ecg'])
                            if 'DIR_ecg' in point:
                                ecg_dir.append(point['DIR_ecg'])
                else:
                    # Fallback para dicionário
                    angles_esq = raw_data.get('ESQ_angle', [])
                    angles_dir = raw_data.get('DIR_angle', [])
                    emg_esq = raw_data.get('ESQ_emg', [])
                    emg_dir = raw_data.get('DIR_emg', [])
                    ecg_esq = raw_data.get('ESQ_ecg', [])
                    ecg_dir = raw_data.get('DIR_ecg', [])
                
                # Converter para arrays numpy
                angles_esq = np.array(angles_esq, dtype=float)
                angles_dir = np.array(angles_dir, dtype=float)
                emg_esq = np.array(emg_esq, dtype=float)
                emg_dir = np.array(emg_dir, dtype=float)
                ecg_esq = np.array(ecg_esq, dtype=float)
                ecg_dir = np.array(ecg_dir, dtype=float)
                
                # Calcular deltas para cada variável
                delta_angle_esq = np.max(angles_esq) - np.min(angles_esq) if len(angles_esq) > 0 else np.nan
                delta_angle_dir = np.max(angles_dir) - np.min(angles_dir) if len(angles_dir) > 0 else np.nan
                
                delta_emg_esq = np.max(emg_esq) - np.min(emg_esq) if len(emg_esq) > 0 else np.nan
                delta_emg_dir = np.max(emg_dir) - np.min(emg_dir) if len(emg_dir) > 0 else np.nan
                
                delta_ecg_esq = np.max(ecg_esq) - np.min(ecg_esq) if len(ecg_esq) > 0 else np.nan
                delta_ecg_dir = np.max(ecg_dir) - np.min(ecg_dir) if len(ecg_dir) > 0 else np.nan
                
                # Armazenar deltas
                if not np.isnan(delta_angle_esq) and not np.isnan(delta_angle_dir):
                    self.deltas['angulo']['esq'].append(delta_angle_esq)
                    self.deltas['angulo']['dir'].append(delta_angle_dir)
                
                if not np.isnan(delta_emg_esq) and not np.isnan(delta_emg_dir):
                    self.deltas['emg']['esq'].append(delta_emg_esq)
                    self.deltas['emg']['dir'].append(delta_emg_dir)
                
                if not np.isnan(delta_ecg_esq) and not np.isnan(delta_ecg_dir):
                    self.deltas['ecg']['esq'].append(delta_ecg_esq)
                    self.deltas['ecg']['dir'].append(delta_ecg_dir)
                
                print(f"  Sessão {session_id}:")
                print(f"    - Ângulo   | ESQ: {delta_angle_esq:>10.4f}° | DIR: {delta_angle_dir:>10.4f}°")
                print(f"    - EMG      | ESQ: {delta_emg_esq:>10.4f}   | DIR: {delta_emg_dir:>10.4f}")
                print(f"    - ECG      | ESQ: {delta_ecg_esq:>10.4f}   | DIR: {delta_ecg_dir:>10.4f}")
                
            except Exception as e:
                print(f"  Sessão {session_id}: ✗ Erro ao calcular deltas: {e}")
        
        print(f"\n✓ Deltas calculados para as variáveis")
        return True
    
    def calculate_cohens_d(self, group1, group2):
        """
        Calcula o tamanho do efeito Cohen's d para amostras pareadas
        
        Args:
            group1: Array com valores da perna esquerda
            group2: Array com valores da perna direita
            
        Returns:
            float: Cohen's d
        """
        # Para amostras pareadas, calcula d usando as diferenças
        differences = np.array(group1) - np.array(group2)
        mean_diff = np.mean(differences)
        std_diff = np.std(differences, ddof=1)
        
        if std_diff == 0:
            return 0
        
        cohens_d = mean_diff / std_diff
        return cohens_d
    
    def calculate_ci_95(self, data):
        """
        Calcula intervalo de confiança de 95%
        
        Args:
            data: Array de dados
            
        Returns:
            tuple: (lower_bound, upper_bound)
        """
        data_array = np.array(data)
        mean = np.mean(data_array)
        std_err = stats.sem(data_array)
        ci = std_err * stats.t.ppf((1 + 0.95) / 2, len(data_array) - 1)
        
        return (mean - ci, mean + ci)
    
    def perform_paired_ttests(self):
        """
        Realiza testes t pareados para cada variável
        """
        print(f"\n{'='*80}")
        print(f"TESTES T PAREADOS (ESQ vs DIR)")
        print(f"{'='*80}\n")
        
        variables = ['angulo', 'emg', 'ecg']
        variable_names = {
            'angulo': 'Ângulo (Angle)',
            'emg': 'Eletromiografia (EMG)',
            'ecg': 'Eletrocardiografia (ECG)'
        }
        
        for var in variables:
            esq_data = self.deltas[var]['esq']
            dir_data = self.deltas[var]['dir']
            
            if len(esq_data) != len(dir_data) or len(esq_data) < 2:
                print(f"\n✗ {variable_names[var]}: Dados insuficientes ou desemparelhados")
                self.results[var] = None
                continue
            
            # Realizar teste t pareado
            t_stat, p_value = stats.ttest_rel(esq_data, dir_data)
            
            # Calcular diferenças
            differences = np.array(esq_data) - np.array(dir_data)
            mean_diff = np.mean(differences)
            std_diff = np.std(differences, ddof=1)
            std_err_diff = stats.sem(differences)
            
            # Calcular Cohen's d
            cohens_d = self.calculate_cohens_d(esq_data, dir_data)
            
            # Calcular IC 95%
            ci_lower, ci_upper = self.calculate_ci_95(differences)
            
            # Interpretação do p-value
            significance = "✓ SIGNIFICANTE" if p_value < 0.05 else "✗ NÃO SIGNIFICANTE"
            
            # Interpretação do Cohen's d
            if abs(cohens_d) < 0.2:
                effect_size = "Negligenciável"
            elif abs(cohens_d) < 0.5:
                effect_size = "Pequeno"
            elif abs(cohens_d) < 0.8:
                effect_size = "Médio"
            else:
                effect_size = "Grande"
            
            result_dict = {
                'variable': var,
                'variable_name': variable_names[var],
                'n': len(esq_data),
                'esq_data': esq_data,
                'dir_data': dir_data,
                'differences': differences.tolist(),
                't_statistic': t_stat,
                'p_value': p_value,
                'mean_esq': np.mean(esq_data),
                'mean_dir': np.mean(dir_data),
                'mean_diff': mean_diff,
                'std_diff': std_diff,
                'std_err_diff': std_err_diff,
                'ci_lower': ci_lower,
                'ci_upper': ci_upper,
                'cohens_d': cohens_d,
                'effect_size': effect_size,
                'significant': p_value < 0.05
            }
            
            self.results[var] = result_dict
            
            # Imprimir resultados formatados
            print(f"\n{variable_names[var].upper()}")
            print(f"{'-' * 76}")
            print(f"  Estatística t:                {t_stat:>15.6f}")
            print(f"  P-value:                      {p_value:>15.6f} {significance}")
            print(f"  N (pares):                    {len(esq_data):>15}")
            print(f"\n  Média ESQ (parética):         {np.mean(esq_data):>15.6f}")
            print(f"  Média DIR (controle):         {np.mean(dir_data):>15.6f}")
            print(f"  Diferença de médias:          {mean_diff:>15.6f}")
            print(f"  Desvio padrão (diferenças):   {std_diff:>15.6f}")
            print(f"  Erro padrão:                  {std_err_diff:>15.6f}")
            print(f"  IC 95%: [{ci_lower:>10.6f}, {ci_upper:>10.6f}]")
            print(f"  Cohen's d:                    {cohens_d:>15.6f} ({effect_size})")
    
    def generate_excel_output(self, output_file="ttest_pareado_resultados.xlsx"):
        """
        Gera arquivo Excel com resultados
        
        Args:
            output_file: Nome do arquivo Excel de saída
        """
        print(f"\n{'='*80}")
        print(f"GERANDO ARQUIVO EXCEL")
        print(f"{'='*80}\n")
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("✗ openpyxl não instalado. Pulando geração de Excel.")
            return
        
        output_path = Path(output_file)
        
        # Criar workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remover sheet padrão
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        significant_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        nonsignificant_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Resumo dos resultados
        ws_summary = wb.create_sheet("Resumo")
        
        ws_summary['A1'] = "TESTE T PAREADO - RESUMO DOS RESULTADOS"
        ws_summary['A1'].font = Font(bold=True, size=14)
        ws_summary.merge_cells('A1:J1')
        
        # Cabeçalhos
        headers = ['Variável', 'N', 'Média ESQ', 'Média DIR', 'Diferença Média', 
                   'Desvio Padrão', 'Teste-t', 'P-value', "Cohen's d", 'Significância']
        
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Dados
        row = 4
        for var_key, result in self.results.items():
            if result is None:
                continue
            
            ws_summary.cell(row=row, column=1).value = result['variable_name']
            ws_summary.cell(row=row, column=2).value = result['n']
            ws_summary.cell(row=row, column=3).value = round(result['mean_esq'], 6)
            ws_summary.cell(row=row, column=4).value = round(result['mean_dir'], 6)
            ws_summary.cell(row=row, column=5).value = round(result['mean_diff'], 6)
            ws_summary.cell(row=row, column=6).value = round(result['std_diff'], 6)
            ws_summary.cell(row=row, column=7).value = round(result['t_statistic'], 6)
            ws_summary.cell(row=row, column=8).value = round(result['p_value'], 6)
            ws_summary.cell(row=row, column=9).value = round(result['cohens_d'], 6)
            ws_summary.cell(row=row, column=10).value = "SIM" if result['significant'] else "NÃO"
            
            # Colorir linha baseado em significância
            fill_color = significant_fill if result['significant'] else nonsignificant_fill
            for col in range(1, 11):
                ws_summary.cell(row=row, column=col).fill = fill_color
                ws_summary.cell(row=row, column=col).border = border
            
            row += 1
        
        # Ajustar largura das colunas
        ws_summary.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws_summary.column_dimensions[col].width = 15
        
        # Sheet 2: Dados detalhados para cada variável
        for var_key, result in self.results.items():
            if result is None:
                continue
            
            ws = wb.create_sheet(result['variable_name'][:31])  # Excel limita nome a 31 caracteres
            
            # Título
            ws['A1'] = f"DETALHES - {result['variable_name'].upper()}"
            ws['A1'].font = Font(bold=True, size=12)
            ws.merge_cells('A1:D1')
            
            # Informações gerais
            ws['A3'] = "Informações Gerais"
            ws['A3'].font = Font(bold=True)
            
            general_info = [
                ['Número de pares:', result['n']],
                ['Estatística t:', round(result['t_statistic'], 6)],
                ['P-value:', round(result['p_value'], 6)],
                ['Cohen\'s d:', round(result['cohens_d'], 6)],
                ['Tamanho do Efeito:', result['effect_size']],
                ['Significante (α=0.05)?', 'Sim' if result['significant'] else 'Não']
            ]
            
            for idx, (label, value) in enumerate(general_info, 4):
                ws[f'A{idx}'] = label
                ws[f'B{idx}'] = value
                ws[f'A{idx}'].font = Font(bold=True)
            
            # Estatísticas das diferenças
            ws['A11'] = "Estatísticas das Diferenças"
            ws['A11'].font = Font(bold=True)
            
            diff_stats = [
                ['Média das diferenças:', round(result['mean_diff'], 6)],
                ['Desvio padrão:', round(result['std_diff'], 6)],
                ['Erro padrão:', round(result['std_err_diff'], 6)],
                ['IC 95% Inferior:', round(result['ci_lower'], 6)],
                ['IC 95% Superior:', round(result['ci_upper'], 6)]
            ]
            
            for idx, (label, value) in enumerate(diff_stats, 12):
                ws[f'A{idx}'] = label
                ws[f'B{idx}'] = value
                ws[f'A{idx}'].font = Font(bold=True)
            
            # Dados brutos
            ws['A18'] = "Dados Brutos"
            ws['A18'].font = Font(bold=True)
            
            ws['A19'] = "Sessão"
            ws['B19'] = "ESQ (Parética)"
            ws['C19'] = "DIR (Controle)"
            ws['D19'] = "Diferença"
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}19'].fill = header_fill
                ws[f'{col}19'].font = header_font
                ws[f'{col}19'].border = border
            
            for idx, (esq, dir_val) in enumerate(zip(result['esq_data'], result['dir_data']), 20):
                ws[f'A{idx}'] = idx - 19 + 18  # ID da sessão
                ws[f'B{idx}'] = round(esq, 6)
                ws[f'C{idx}'] = round(dir_val, 6)
                ws[f'D{idx}'] = round(esq - dir_val, 6)
                
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{idx}'].border = border
        
        # Salvar workbook
        wb.save(output_path)
        print(f"✓ Arquivo Excel gerado: {output_path}")
    
    def generate_csv_output(self, output_file="ttest_pareado_resultados.csv"):
        """
        Gera arquivo CSV com resultados
        
        Args:
            output_file: Nome do arquivo CSV de saída
        """
        print(f"\n{'='*80}")
        print(f"GERANDO ARQUIVO CSV")
        print(f"{'='*80}\n")
        
        output_path = Path(output_file)
        
        csv_data = []
        csv_data.append(['TESTE T PAREADO - RESULTADOS', '', '', '', '', '', '', '', ''])
        csv_data.append(['Data da análise', datetime.now().strftime("%Y-%m-%d %H:%M:%S"), '', '', '', '', '', '', ''])
        csv_data.append(['Sessões analisadas', f"{self.session_ids[0]}-{self.session_ids[-1]}", '', '', '', '', '', '', ''])
        csv_data.append(['', '', '', '', '', '', '', '', ''])
        
        csv_data.append(['Variável', 'N', 'Média ESQ', 'Média DIR', 'Diferença Média', 
                        'Desvio Padrão', 'Teste-t', 'P-value', "Cohen's d", 'Significância', 
                        'IC 95% Inferior', 'IC 95% Superior', 'Tamanho do Efeito', 'Interpretação'])
        
        for var_key, result in self.results.items():
            if result is None:
                continue
            
            # Interpretação
            if result['significant']:
                if result['p_value'] < 0.001:
                    interpretation = "Diferença MUITO significante entre pernas (p < 0.001)"
                else:
                    interpretation = f"Diferença significante entre pernas (p = {result['p_value']:.4f})"
            else:
                interpretation = f"Sem diferença significante (p = {result['p_value']:.4f})"
            
            csv_data.append([
                result['variable_name'],
                result['n'],
                round(result['mean_esq'], 6),
                round(result['mean_dir'], 6),
                round(result['mean_diff'], 6),
                round(result['std_diff'], 6),
                round(result['t_statistic'], 6),
                round(result['p_value'], 6),
                round(result['cohens_d'], 6),
                "SIM" if result['significant'] else "NÃO",
                round(result['ci_lower'], 6),
                round(result['ci_upper'], 6),
                result['effect_size'],
                interpretation
            ])
        
        # Salvar CSV
        df_csv = pd.DataFrame(csv_data)
        df_csv.to_csv(output_path, index=False, header=False, encoding='utf-8')
        print(f"✓ Arquivo CSV gerado: {output_path}")
    
    def print_formatted_output(self):
        """
        Imprime saída formatada com interpretações
        """
        print(f"\n{'='*80}")
        print(f"INTERPRETAÇÃO DOS RESULTADOS")
        print(f"{'='*80}\n")
        
        for var_key, result in self.results.items():
            if result is None:
                print(f"\n✗ {var_key.upper()}: Dados insuficientes")
                continue
            
            print(f"\n{'─' * 76}")
            print(f"VARIÁVEL: {result['variable_name'].upper()}")
            print(f"{'─' * 76}\n")
            
            # Resumo estatístico
            print(f"  Resumo Estatístico:")
            print(f"    • Número de pares observados: {result['n']}")
            print(f"    • Média da perna esquerda (parética): {result['mean_esq']:.6f}")
            print(f"    • Média da perna direita (controle): {result['mean_dir']:.6f}")
            print(f"    • Diferença média: {result['mean_diff']:.6f}")
            print(f"    • Desvio padrão das diferenças: {result['std_diff']:.6f}")
            
            # Teste t pareado
            print(f"\n  Teste T Pareado:")
            print(f"    • Estatística t: {result['t_statistic']:.6f}")
            print(f"    • P-value: {result['p_value']:.6f}")
            print(f"    • Intervalo de Confiança 95%: [{result['ci_lower']:.6f}, {result['ci_upper']:.6f}]")
            
            # Tamanho do efeito
            print(f"\n  Tamanho do Efeito:")
            print(f"    • Cohen's d: {result['cohens_d']:.6f}")
            print(f"    • Classificação: {result['effect_size']}")
            
            # Conclusão
            print(f"\n  Conclusão:")
            if result['significant']:
                print(f"    ✓ HÁ diferença SIGNIFICANTE entre pernas (p < 0.05)")
                print(f"    • A perna {'direita' if result['mean_diff'] < 0 else 'esquerda'} apresenta valores {'maiores' if result['mean_diff'] > 0 else 'menores'}")
                print(f"    • O tamanho do efeito é {result['effect_size'].lower()}")
            else:
                print(f"    ✗ NÃO há diferença significante entre pernas (p ≥ 0.05)")
                print(f"    • As pernas apresentam padrões similares nesta variável")
    
    def run_analysis(self):
        """
        Executa a análise completa
        """
        try:
            print("\n" + "=" * 80)
            print("ANÁLISE DE TESTE T PAREADO - COMPARAÇÃO PERNA ESQUERDA vs DIREITA")
            print("=" * 80)
            
            self.connect_db()
            self.extract_session_data()
            self.calculate_deltas()
            self.perform_paired_ttests()
            self.print_formatted_output()
            self.generate_excel_output()
            self.generate_csv_output()
            
            print(f"\n{'='*80}")
            print("✓ ANÁLISE CONCLUÍDA COM SUCESSO")
            print(f"{'='*80}\n")
            
        except Exception as e:
            print(f"\n✗ ERRO DURANTE A ANÁLISE: {e}")
            raise
        finally:
            self.close_db()


if __name__ == "__main__":
    analyzer = PairedTTestAnalyzer()
    analyzer.run_analysis()
